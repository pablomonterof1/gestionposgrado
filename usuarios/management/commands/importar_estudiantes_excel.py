from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.contrib.auth.hashers import make_password
from django.db import transaction
from openpyxl import load_workbook

from usuarios.models import PerfilUsuario  # cambia por tu app real


class Command(BaseCommand):
    help = 'Importa estudiantes desde un archivo Excel (.xlsx) en lotes grandes'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta del archivo Excel .xlsx'
        )
        parser.add_argument(
            '--batch-size',
            type=int,
            default=200,
            help='Cantidad de registros por lote (default: 200)'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        batch_size = options['batch_size']

        try:
            wb = load_workbook(filename=archivo, data_only=True, read_only=True)
        except Exception as e:
            raise CommandError(f'No se pudo abrir el archivo Excel: {e}')

        ws = wb.active

        filas = ws.iter_rows(values_only=True)

        try:
            encabezados_raw = next(filas)
        except StopIteration:
            raise CommandError('El archivo está vacío.')

        encabezados = [
            str(valor).strip().lower() if valor is not None else ''
            for valor in encabezados_raw
        ]

        columnas_requeridas = ['nombres', 'apellidos', 'correo', 'cedula']
        for col in columnas_requeridas:
            if col not in encabezados:
                raise CommandError(
                    f'No se encontró la columna requerida "{col}". '
                    f'Encabezados encontrados: {encabezados}'
                )

        idx_nombres = encabezados.index('nombres')
        idx_apellidos = encabezados.index('apellidos')
        idx_correo = encabezados.index('correo')
        idx_cedula = encabezados.index('cedula')

        total_creados = 0
        total_omitidos = 0
        total_errores = 0

        lote = []
        numero_fila = 1

        for row in filas:
            numero_fila += 1
            lote.append((numero_fila, row))

            if len(lote) >= batch_size:
                creados, omitidos, errores = self.procesar_lote(
                    lote,
                    idx_nombres,
                    idx_apellidos,
                    idx_correo,
                    idx_cedula
                )
                total_creados += creados
                total_omitidos += omitidos
                total_errores += errores
                self.stdout.write(
                    self.style.SUCCESS(
                        f'Lote procesado hasta fila {numero_fila}. '
                        f'Creados acumulados: {total_creados}'
                    )
                )
                lote = []

        if lote:
            creados, omitidos, errores = self.procesar_lote(
                lote,
                idx_nombres,
                idx_apellidos,
                idx_correo,
                idx_cedula
            )
            total_creados += creados
            total_omitidos += omitidos
            total_errores += errores

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'Total creados: {total_creados}'))
        self.stdout.write(self.style.WARNING(f'Total omitidos: {total_omitidos}'))
        self.stdout.write(self.style.ERROR(f'Total errores: {total_errores}'))

    def procesar_lote(self, lote, idx_nombres, idx_apellidos, idx_correo, idx_cedula):
        omitidos = 0
        errores = 0

        datos_validos = []
        cedulas_lote = set()
        correos_lote = set()

        for fila, row in lote:
            try:
                nombres = row[idx_nombres] if idx_nombres < len(row) else None
                apellidos = row[idx_apellidos] if idx_apellidos < len(row) else None
                correo = row[idx_correo] if idx_correo < len(row) else None
                cedula = row[idx_cedula] if idx_cedula < len(row) else None

                nombres = str(nombres).strip() if nombres is not None else ''
                apellidos = str(apellidos).strip() if apellidos is not None else ''
                correo = str(correo).strip().lower() if correo is not None else ''
                cedula = str(cedula).strip() if cedula is not None else ''

                if cedula:
                    cedula = cedula.zfill(10)

                if not nombres or not apellidos or not correo or not cedula:
                    omitidos += 1
                    continue

                if not cedula.isdigit() or len(cedula) != 10:
                    omitidos += 1
                    continue

                try:
                    validate_email(correo)
                except ValidationError:
                    omitidos += 1
                    continue

                if cedula in cedulas_lote:
                    omitidos += 1
                    continue

                if correo in correos_lote:
                    omitidos += 1
                    continue

                datos_validos.append({
                    'nombres': nombres,
                    'apellidos': apellidos,
                    'correo': correo,
                    'cedula': cedula,
                })

                cedulas_lote.add(cedula)
                correos_lote.add(correo)

            except Exception:
                errores += 1

        if not datos_validos:
            return 0, omitidos, errores

        cedulas_consulta = [d['cedula'] for d in datos_validos]
        correos_consulta = [d['correo'] for d in datos_validos]

        usernames_existentes = set(
            User.objects.filter(username__in=cedulas_consulta)
            .values_list('username', flat=True)
        )

        correos_existentes = set(
            correo.strip().lower()
            for correo in User.objects.filter(email__in=correos_consulta)
            .values_list('email', flat=True)
            if correo
        )

        cedulas_perfil_existentes = set(
            PerfilUsuario.objects.filter(ci__in=cedulas_consulta)
            .values_list('ci', flat=True)
        )

        usuarios_a_crear = []
        perfiles_data = []

        for item in datos_validos:
            if item['cedula'] in usernames_existentes:
                omitidos += 1
                continue

            if item['correo'] in correos_existentes:
                omitidos += 1
                continue

            if item['cedula'] in cedulas_perfil_existentes:
                omitidos += 1
                continue

            usuarios_a_crear.append(
                User(
                    username=item['cedula'],
                    password=make_password(item['cedula']),
                    first_name=item['nombres'],
                    last_name=item['apellidos'],
                    email=item['correo']
                )
            )
            perfiles_data.append(item)

        if not usuarios_a_crear:
            return 0, omitidos, errores

        try:
            with transaction.atomic():
                User.objects.bulk_create(usuarios_a_crear, batch_size=200)

                usuarios_creados = User.objects.filter(
                    username__in=[u.username for u in usuarios_a_crear]
                ).only('id', 'username')

                mapa_usuarios = {u.username: u for u in usuarios_creados}

                perfiles_a_crear = []
                for item in perfiles_data:
                    user = mapa_usuarios.get(item['cedula'])
                    if user:
                        perfiles_a_crear.append(
                            PerfilUsuario(
                                user=user,
                                rol=1,
                                ci=item['cedula']
                            )
                        )
                    else:
                        errores += 1

                if perfiles_a_crear:
                    PerfilUsuario.objects.bulk_create(perfiles_a_crear, batch_size=200)

        except Exception as e:
            raise CommandError(f'Error durante la importación del lote: {e}')

        return len(usuarios_a_crear), omitidos, errores